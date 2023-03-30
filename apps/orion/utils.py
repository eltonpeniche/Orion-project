import calendar
import os
import string
from datetime import date, datetime

import convertapi
from django.core.paginator import Paginator
from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side

from apps.orion.models import CargaHoraria

PER_PAGE = os.environ.get('PER_PAGE', 10)
CONVERTEAPI_SECRET_KEY = os.environ.get('CONVERTEAPI_SECRET_KEY')

def paginacao(request, object_list):
    paginator = Paginator(object_list, PER_PAGE) # Show 25 contacts per page.
    page_number = request.GET.get('page')
    return paginator.get_page(page_number)



#verifica se dois horarios de inicio e fim se sobrepoe
def horarios_se_sobrepoe(h1_inicio, h1_fim, h2_inicio, h2_fim):
    if h1_inicio <= h2_inicio < h1_fim or h2_inicio <= h1_inicio < h2_fim:
        return True
    return False

def converter_data(data):
    if len(data.split('/')) == 3:
        dia, mes, ano  = data.split('/')
    elif len(data.split('-')) == 3:
        dia, mes, ano  = data.split('-')
    else:
        return data;
        #raise Exception("Data inserida Invalida")
    
    return f"{ano}-{mes}-{dia}"

 

MODELO_PONTO = "media/ponto_modelo.xlsx"
PONTO = "media/ponto.xlsx"
CAMINHO = 'media'
COLUNAS = list(string.ascii_lowercase)[:9]
LINHA_INICIAL = 15

def num_dias_mes(data = None):
    if data is None:
        now = datetime.datetime.now()
        return calendar.monthrange(now.year, now.month)[1]
    else:
        data = datetime.strptime(data, '%Y-%m')
        return calendar.monthrange(data.year, data.month)[1]

def fill_line(linha, sh):
    # Definir a cor de fundo da célula A1
    color_fill = PatternFill(start_color='F5F5F5',
                        end_color='F5F5F5',
                        fill_type='solid')

    for letra_coluna in COLUNAS:
        celula = str(letra_coluna) + str(linha)
        sh[celula].fill = color_fill

    return sh

def thin_border(linha, sh):
    # Definir a borda ao redor da célula A1
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
   

    for letra_coluna in COLUNAS:
        celula = str(letra_coluna) + str(linha)
        sh[celula].border = thin_border

    return sh

def gerar_relatorio(funcionario, mes_referencia):
    ano, mes = mes_referencia.split('-')
    
    wb = load_workbook(filename=MODELO_PONTO)
    sh = wb.active
    sh['A9'] = f"{funcionario.user.first_name} {funcionario.user.last_name}"
    sh['A11'] = f"{funcionario.get_tipo_display()}"
    sh['I1'] = "Em "+ datetime.now().strftime("%d/%m/%Y")
    sh['A6'] = f"REFERÊNCIA {mes}/{ano}"
    
    mes_trabalhado = date(int(ano), int(mes), 1)

    query_carga_horaria = CargaHoraria.objects.select_related('tecnico').filter(data__month=mes, data__year=ano).filter(tecnico=funcionario.user_id).order_by('data')
    
    cont = 1
    for i in range(LINHA_INICIAL, LINHA_INICIAL + num_dias_mes(mes_referencia)):
        sh['A' + str(i)] = cont 
        cont+=1
        if i % 2 == 0:
            sh = fill_line(i, sh)
        sh = thin_border(i, sh)

    data = mes_trabalhado
    periodo = 0

    linha = LINHA_INICIAL
    for e in query_carga_horaria:
        if e.data != data:
            periodo=1
            razao = e.data - data 
            linha+=razao.days
        else:
            periodo+=1
        
        if periodo == 1:
            sh['B' + str(linha)] = e.hora_inicio  
            sh['C' + str(linha)] = e.hora_termino
        elif periodo == 2: 
            sh['D' + str(linha)] =  e.hora_inicio 
            sh['E' + str(linha)] =  e.hora_termino
        elif periodo == 3:
            sh['F' + str(linha)] = e.hora_inicio
            sh['G' + str(linha)] = e.hora_termino
        #print(e.data, e.hora_inicio,e.hora_termino, periodo)
        
        data = e.data
            
    wb.save(filename='media/ponto.xlsx')


    #Code snippet is using the ConvertAPI Python Client: https://github.com/ConvertAPI/convertapi-python
    convertapi.api_secret = CONVERTEAPI_SECRET_KEY
    convertapi.convert('pdf', {
        'File': PONTO
    }, from_format = 'xlsx').save_files(CAMINHO)